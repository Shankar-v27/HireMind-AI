import csv
import io
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.security import get_password_hash
from app.db.session import get_db
from app.models.core import (
    Candidate,
    Company,
    GDSession,
    Interview,
    InterviewSession,
    InterviewCandidate,
    ProctoringEvent,
    Question,
    Response,
    Round,
    RoundSession,
    Strike,
    Verification,
)
from app.models.user import User, UserRole, role_value
from app.routers.auth import get_current_admin
from app.schemas.core import CompanyCreate, CompanyRead, BulkCompanyResult


router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/stats")
def admin_stats(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
) -> dict:
    """Return counts for admin dashboard."""
    total_companies = db.query(Company).count()
    total_interviews = db.query(Interview).count()
    total_candidates = db.query(Candidate).count()
    return {"total_companies": total_companies, "total_interviews": total_interviews, "total_candidates": total_candidates}


@router.post("/companies", response_model=CompanyRead)
def create_company(
    payload: CompanyCreate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
) -> CompanyRead:
    existing_user = db.query(User).filter(User.email == payload.admin_email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Admin email already in use")

    admin_user = User(
        email=payload.admin_email,
        full_name=payload.admin_full_name,
        hashed_password=get_password_hash(payload.admin_password),
        role=role_value(UserRole.COMPANY),
    )
    db.add(admin_user)
    db.flush()

    company = Company(
        name=payload.name,
        contact_email=payload.contact_email or payload.admin_email,
        user_id=admin_user.id,
    )
    db.add(company)
    db.commit()
    db.refresh(company)
    return company


@router.post("/cleanup-orphan-candidate-users")
def cleanup_orphan_candidate_users(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
) -> dict:
    """Remove users with role=candidate who have no Candidate row (e.g. left after company was deleted)."""
    candidate_role = role_value(UserRole.CANDIDATE)
    candidate_user_ids = {c.user_id for c in db.query(Candidate).all()}
    q = db.query(User).filter(User.role == candidate_role)
    if candidate_user_ids:
        q = q.filter(~User.id.in_(candidate_user_ids))
    orphan_users = q.all()
    deleted = 0
    for u in orphan_users:
        db.delete(u)
        deleted += 1
    db.commit()
    return {"ok": True, "deleted": deleted, "message": f"Removed {deleted} orphan candidate user(s). Those emails can be used again."}


@router.get("/companies", response_model=list[CompanyRead])
def list_companies(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
) -> list[CompanyRead]:
    companies = db.query(Company).order_by(Company.created_at.desc()).all()
    return companies


@router.delete("/companies/{company_id}")
def delete_company(
    company_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
) -> dict:
    """Delete company and all associated data from DB (interviews, candidates, responses, etc.)."""
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    admin_user_id = company.user_id
    interview_ids = [i.id for i in db.query(Interview).filter(Interview.company_id == company_id).all()]
    round_ids = [r.id for r in db.query(Round).filter(Round.interview_id.in_(interview_ids)).all()] if interview_ids else []
    candidates = db.query(Candidate).filter(Candidate.company_id == company_id).all()
    candidate_ids = [c.id for c in candidates]
    candidate_user_ids = [c.user_id for c in candidates]

    if candidate_ids:
        db.query(Response).filter(Response.candidate_id.in_(candidate_ids)).delete(synchronize_session=False)
        db.query(Strike).filter(Strike.candidate_id.in_(candidate_ids)).delete(synchronize_session=False)
        db.query(ProctoringEvent).filter(ProctoringEvent.candidate_id.in_(candidate_ids)).delete(synchronize_session=False)
        db.query(Verification).filter(Verification.candidate_id.in_(candidate_ids)).delete(synchronize_session=False)
        db.query(RoundSession).filter(RoundSession.candidate_id.in_(candidate_ids)).delete(synchronize_session=False)
        db.query(InterviewSession).filter(InterviewSession.candidate_id.in_(candidate_ids)).delete(synchronize_session=False)
    if interview_ids:
        db.query(InterviewCandidate).filter(InterviewCandidate.interview_id.in_(interview_ids)).delete(synchronize_session=False)
    if round_ids:
        db.query(GDSession).filter(GDSession.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(RoundSession).filter(RoundSession.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(InterviewSession).filter(InterviewSession.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(Question).filter(Question.round_id.in_(round_ids)).delete(synchronize_session=False)
    if interview_ids:
        db.query(Round).filter(Round.interview_id.in_(interview_ids)).delete(synchronize_session=False)
        db.query(Interview).filter(Interview.company_id == company_id).delete(synchronize_session=False)
    db.query(Candidate).filter(Candidate.company_id == company_id).delete(synchronize_session=False)
    db.delete(company)
    for uid in candidate_user_ids:
        u = db.query(User).filter(User.id == uid).first()
        if u:
            db.delete(u)
    admin_user = db.query(User).filter(User.id == admin_user_id).first()
    if admin_user:
        db.delete(admin_user)
    db.commit()
    return {"ok": True, "message": "Company and all associated data deleted"}


def _parse_rows_from_upload(file: UploadFile) -> list[dict[str, Any]]:
    content = file.file.read()
    file.file.seek(0)
    rows: list[dict[str, Any]] = []
    fn = (file.filename or "").lower()
    if fn.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
        for r in reader:
            rows.append(dict(r))
    elif fn.endswith(".xlsx") or fn.endswith(".xls"):
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            return []
        header = [str(c.value or "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2):
            rows.append({header[i]: str(c.value or "").strip() for i, c in enumerate(row) if i < len(header)})
        wb.close()
    return rows


@router.post("/companies/bulk", response_model=BulkCompanyResult)
def bulk_create_companies(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
) -> BulkCompanyResult:
    # Expected columns: name, admin_email, password (or admin_password)
    rows = _parse_rows_from_upload(file)
    created = 0
    errors: list[dict] = []
    for i, row in enumerate(rows):
        row_num = i + 2
        name = (row.get("name") or "").strip()
        admin_email = (row.get("admin_email") or row.get("email") or "").strip().lower()
        password = (row.get("password") or row.get("admin_password") or "").strip()
        if not name or not admin_email or not password:
            errors.append({"row": row_num, "error": "Missing name, admin_email, or password"})
            continue
        if db.query(User).filter(User.email == admin_email).first():
            errors.append({"row": row_num, "error": f"Email already in use: {admin_email}"})
            continue
        if db.query(Company).filter(Company.name == name).first():
            errors.append({"row": row_num, "error": f"Company name already exists: {name}"})
            continue
        try:
            user = User(
                email=admin_email,
                full_name=(row.get("admin_full_name") or row.get("full_name") or "").strip() or None,
                hashed_password=get_password_hash(password),
                role=role_value(UserRole.COMPANY),
            )
            db.add(user)
            db.flush()
            company = Company(name=name, contact_email=admin_email, user_id=user.id)
            db.add(company)
            db.commit()
            created += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": row_num, "error": str(e)})
    return BulkCompanyResult(created=created, failed=len(errors), errors=errors)

