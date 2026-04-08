from datetime import datetime
import csv
import io
import logging
import re
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.security import get_password_hash
from app.db.session import get_db
from app.models.core import (
    Candidate, Company, GDSession, Interview, InterviewCandidate,
    InterviewSession, ProctoringEvent, Question, Response,
    Round, RoundSession, Strike, Verification,
)
from app.models.user import User, UserRole, role_value
from app.routers.auth import get_current_company
from app.schemas.core import (
    BulkCandidateResult,
    BulkQuestionsResult,
    CandidateCreate,
    CandidatePerformanceSummary,
    CandidateRoundSummary,
    CandidateRead,
    CompanyRead,
    EnrolledCandidateRead,
    GenerateQuestionsRequest,
    InterviewCreate,
    InterviewRead,
    InterviewUpdate,
    QuestionCreate,
    QuestionRead,
    QuestionUpdate,
    ResponseWithCandidate,
    StructuredResponseDetail,
    LiveAssistRequest,
    LiveAssistResponse,
    LiveInterviewScoreRequest,
    LiveInterviewScoreResponse,
    LiveInterviewStartRequest,
    LiveInterviewStartResponse,
    RoundConfigUpdate,
    RoundCreate,
    RoundRead,
    RoundUpdate,
    RoundsReorderRequest,
    TopPerformerEntry,
    VerificationRead,
)

logger = logging.getLogger(__name__)


router = APIRouter(prefix="/company", tags=["company"])


def _extract_option_text(question: Question | None, option_key: str | None) -> str | None:
    if not question or not option_key or not question.options:
        return None
    if isinstance(question.options, dict):
        value = question.options.get(option_key) or question.options.get(option_key.upper()) or question.options.get(option_key.lower())
        return str(value) if value is not None else None
    return None


def _normalize_answer_letter(answer: str | None) -> str | None:
    if not answer:
        return None
    match = re.match(r"\s*([A-Z])\b", answer.strip(), re.IGNORECASE)
    return match.group(1).upper() if match else answer.strip()


def _has_plagiarism_warning(flags: dict[str, Any] | None) -> bool:
    if not flags:
        return False
    plagiarism = flags.get("plagiarism") or {}
    cross = flags.get("cross_plagiarism") or {}
    return bool(plagiarism.get("warning") or cross.get("warning"))


def _normalize_status_label(status: str) -> str:
    return status.replace("_", " ").title()


def _get_recommendation_status(
    overall_score: float,
    rank: int | None,
    shortlist_count: int | None,
) -> str:
    if shortlist_count and shortlist_count > 0:
        if rank is not None and rank <= shortlist_count:
            return "recommended" if overall_score >= 80 else "rejected"
        if overall_score >= 50:
            return "waiting_list"
        return "rejected"
    if overall_score > 75:
        return "recommended"
    if overall_score >= 50:
        return "waiting_list"
    return "rejected"


def _build_suitability_comment(
    candidate_name: str | None,
    recommendation_status: str,
    overall_score: float,
    total_warnings: int,
) -> str:
    name = candidate_name or "This candidate"
    if recommendation_status == "recommended":
        base = f"{name} suits the role based on the current interview performance."
    elif recommendation_status == "waiting_list":
        base = f"{name} shows partial fit for the role but needs closer recruiter review."
    else:
        base = f"{name} does not currently suit the role based on the interview performance."
    if total_warnings > 0:
        return f"{base} Overall score is {overall_score:.2f}/100 with {total_warnings} warning(s), so recruiter review should also consider interview discipline."
    return f"{base} Overall score is {overall_score:.2f}/100."


def _make_json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _make_json_safe(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_make_json_safe(item) for item in value]
    return value


def _sanitize_ocr_data(ocr_data: dict[str, Any] | None) -> dict[str, Any] | None:
    data = dict(ocr_data or {})
    data.pop("reference_photo_base64", None)
    data.pop("reference_id_base64", None)
    return data or None


def get_company_for_user(db: Session, user: User) -> Company:
    company = db.query(Company).filter(Company.user_id == user.id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found for user")
    return company


def _recalculate_round_session_score(db: Session, round_id: int, candidate_id: int) -> None:
    session = db.query(RoundSession).filter(
        RoundSession.round_id == round_id,
        RoundSession.candidate_id == candidate_id,
    ).first()
    if not session:
        return
    responses = db.query(Response).filter(
        Response.round_id == round_id,
        Response.candidate_id == candidate_id,
    ).all()
    session.total_score = round(sum(float(r.score or 0) for r in responses), 2)


def _compute_interview_rankings(db: Session, interview: Interview) -> list[dict[str, Any]]:
    rounds = db.query(Round).filter(Round.interview_id == interview.id).order_by(Round.order).all()
    links = db.query(InterviewCandidate).filter(InterviewCandidate.interview_id == interview.id).all()
    candidate_ids = [l.candidate_id for l in links]
    if not candidate_ids:
        return []

    scores: dict[int, dict[str, Any]] = {cid: {"total": 0.0, "rounds": {}} for cid in candidate_ids}

    for rnd in rounds:
        weight = rnd.weightage / 100.0 if rnd.weightage else 0
        sessions = db.query(RoundSession).filter(
            RoundSession.round_id == rnd.id,
            RoundSession.candidate_id.in_(candidate_ids),
        ).all()
        question_map: dict[int, Question] = {}
        if rnd.type == "CODING":
            q_list = db.query(Question).filter(Question.round_id == rnd.id).all()
            question_map = {q.id: q for q in q_list}
        for sess in sessions:
            effective_total = sess.total_score
            effective_max = sess.max_possible_score
            if rnd.type == "CODING":
                responses = db.query(Response).filter(
                    Response.round_id == rnd.id,
                    Response.candidate_id == sess.candidate_id,
                ).all()
                effective_total = round(
                    sum(0.0 if _has_plagiarism_warning(r.flags) else float(r.score or 0) for r in responses),
                    2,
                )
                if responses:
                    effective_max = sum(float(question_map.get(r.question_id).max_score or 0) for r in responses if question_map.get(r.question_id)) or effective_max
            if effective_total is not None and effective_max:
                norm = (effective_total / effective_max) * 100
            elif effective_total is not None:
                norm = effective_total
            else:
                norm = 0
            weighted = norm * weight
            scores[sess.candidate_id]["total"] += weighted
            scores[sess.candidate_id]["rounds"][rnd.id] = {
                "normalized_score": round(norm, 2),
                "weighted_score": round(weighted, 2),
                "raw_score": effective_total,
                "max_score": effective_max,
                "status": sess.status,
                "weightage": rnd.weightage or 0,
                "notes": sess.notes,
            }

    ranking = []
    for cid in candidate_ids:
        candidate = db.query(Candidate).filter(Candidate.id == cid).first()
        if not candidate:
            continue
        ranking.append({
            "candidate_id": cid,
            "candidate_email": candidate.user.email,
            "candidate_name": candidate.user.full_name,
            "total_weighted_score": round(scores[cid]["total"], 2),
            "round_scores": scores[cid]["rounds"],
        })
    ranking.sort(key=lambda item: item["total_weighted_score"], reverse=True)
    for index, item in enumerate(ranking, start=1):
        item["rank"] = index
    for item in ranking:
        item["recommendation_status"] = _get_recommendation_status(
            item["total_weighted_score"],
            item.get("rank"),
            interview.shortlist_count,
        )
    return ranking


def _build_candidate_performance_summary(
    db: Session,
    interview: Interview,
    candidate: Candidate,
    verification: Verification | None,
) -> CandidatePerformanceSummary:
    rankings = _compute_interview_rankings(db, interview)
    rank_entry = next((item for item in rankings if item["candidate_id"] == candidate.id), None)
    rounds = db.query(Round).filter(Round.interview_id == interview.id).order_by(Round.order).all()
    strikes = db.query(Strike).filter(
        Strike.interview_id == interview.id,
        Strike.candidate_id == candidate.id,
    ).all()
    strike_map = {strike.round_id: strike.strikes for strike in strikes}
    total_warnings = sum(strike.strikes for strike in strikes)
    overall_score = rank_entry["total_weighted_score"] if rank_entry else 0
    recommendation_status = _get_recommendation_status(
        overall_score,
        rank_entry["rank"] if rank_entry else None,
        interview.shortlist_count,
    )

    round_summaries: list[CandidateRoundSummary] = []
    for rnd in rounds:
        sess = db.query(RoundSession).filter(
            RoundSession.round_id == rnd.id,
            RoundSession.candidate_id == candidate.id,
        ).first()
        ranking_details = (rank_entry or {}).get("round_scores", {}).get(rnd.id, {})
        round_summaries.append(
            CandidateRoundSummary(
                round_id=rnd.id,
                round_type=rnd.type,
                round_order=rnd.order,
                status=sess.status if sess else "not_started",
                score=ranking_details.get("raw_score", sess.total_score if sess else None),
                max_score=ranking_details.get("max_score", sess.max_possible_score if sess else None),
                normalized_score=ranking_details.get("normalized_score"),
                weighted_score=ranking_details.get("weighted_score"),
                weightage=rnd.weightage or 0,
                notes=sess.notes if sess else None,
                warning_count=strike_map.get(rnd.id, 0),
            )
        )

    def _avg(types: set[str]) -> float:
        values = [
            round_item.normalized_score
            for round_item in round_summaries
            if round_item.round_type in types and round_item.normalized_score is not None
        ]
        return round(sum(values) / len(values), 2) if values else 0.0

    score_breakdown = {
        "technical": _avg({"TECH_INTERVIEW", "CODING", "APT_TECH"}),
        "problem_solving": _avg({"CODING", "APT_QUANT", "APT_MIXED"}),
        "communication": _avg({"HR_INTERVIEW", "LIVE_INTERVIEW", "GD"}),
        "overall_fit": round(overall_score, 2),
    }

    question_map = {
        q.id: q
        for q in db.query(Question).join(Round, Question.round_id == Round.id).filter(Round.interview_id == interview.id).all()
    }
    round_map = {rnd.id: rnd for rnd in rounds}
    responses = db.query(Response).filter(
        Response.candidate_id == candidate.id,
        Response.interview_id == interview.id,
    ).order_by(Response.created_at.asc()).all()

    structured_responses: list[StructuredResponseDetail] = []
    for resp in responses:
        question = question_map.get(resp.question_id)
        if not question:
            continue
        round_obj = round_map.get(question.round_id)
        plagiarism_warning = _has_plagiarism_warning(resp.flags)
        selected_option = _normalize_answer_letter(resp.content if question.type == "mcq" else None)
        correct_answer = _normalize_answer_letter(question.correct_answer if question.type == "mcq" else None)
        effective_score = 0 if plagiarism_warning else resp.score
        structured_responses.append(
            StructuredResponseDetail(
                response_id=resp.id,
                round_id=resp.round_id,
                round_type=round_obj.type if round_obj else None,
                question_id=question.id,
                question_content=question.content,
                question_type=question.type,
                candidate_answer=resp.content,
                selected_option=selected_option,
                selected_option_text=_extract_option_text(question, selected_option),
                correct_answer=correct_answer if question.type == "mcq" else question.correct_answer,
                correct_option_text=_extract_option_text(question, correct_answer),
                is_correct=(selected_option == correct_answer) if question.type == "mcq" and selected_option and correct_answer else None,
                score=resp.score,
                effective_score=effective_score,
                max_score=question.max_score,
                grading_method=resp.grading_method,
                grading_details=resp.grading_details,
                plagiarism_warning=plagiarism_warning,
                plagiarism=(resp.flags or {}).get("plagiarism") if resp.flags else None,
                cross_plagiarism=(resp.flags or {}).get("cross_plagiarism") if resp.flags else None,
                created_at=resp.created_at,
            )
        )

    proctoring_events = db.query(ProctoringEvent).filter(
        ProctoringEvent.interview_id == interview.id,
        ProctoringEvent.candidate_id == candidate.id,
    ).order_by(ProctoringEvent.created_at.desc()).all()
    event_rows = [
        {
            "type": event.type,
            "created_at": event.created_at.isoformat(),
            "data": event.data or {},
        }
        for event in proctoring_events
    ]

    return CandidatePerformanceSummary(
        candidate_id=candidate.id,
        candidate_email=candidate.user.email,
        candidate_name=candidate.user.full_name,
        verification_status=verification.status if verification else None,
        overall_rank=rank_entry["rank"] if rank_entry else None,
        total_candidates=len(rankings),
        total_weighted_score=overall_score,
        recommendation_status=recommendation_status,
        suitability_comment=_build_suitability_comment(
            candidate.user.full_name,
            recommendation_status,
            overall_score,
            total_warnings,
        ),
        score_breakdown=score_breakdown,
        total_warnings=total_warnings,
        report="",
        rounds=round_summaries,
        responses=structured_responses,
        verification=VerificationRead(
            id=verification.id,
            candidate_id=verification.candidate_id,
            id_proof_url=verification.id_proof_url,
            photo_url=verification.photo_url,
            resume_url=verification.resume_url,
            status=verification.status,
            ocr_data=_sanitize_ocr_data(verification.ocr_data),
            created_at=verification.created_at,
            updated_at=verification.updated_at,
        ) if verification else None,
        proctoring_events=event_rows,
    )


def _generate_candidate_report_text(
    interview: Interview,
    summary: CandidatePerformanceSummary,
) -> str:
    from app.core.config import get_settings

    settings = get_settings()
    if not settings.claude_api_key:
        return "AI report unavailable (no API key)"

    try:
        import anthropic
        import json

        client = anthropic.Anthropic(api_key=settings.claude_api_key)
        report_payload = {
            "candidate_name": summary.candidate_name,
            "candidate_email": summary.candidate_email,
            "overall_rank": summary.overall_rank,
            "total_candidates": summary.total_candidates,
            "total_weighted_score": summary.total_weighted_score,
            "recommendation_status": summary.recommendation_status,
            "suitability_comment": summary.suitability_comment,
            "score_breakdown": summary.score_breakdown,
            "total_warnings": summary.total_warnings,
            "verification_status": summary.verification_status,
            "rounds": [_make_json_safe(round_item.model_dump()) for round_item in summary.rounds],
            "responses": [_make_json_safe(response.model_dump()) for response in summary.responses],
        }
        prompt = (
            f"Generate a detailed performance evaluation report for candidate {summary.candidate_name or summary.candidate_email} "
            f"in interview '{interview.name}'.\n\n"
            f"Candidate performance data:\n{json.dumps(_make_json_safe(report_payload), indent=2)}\n\n"
            "Requirements:\n"
            "- Explain round-by-round how the candidate performed.\n"
            "- Include a short suitability statement for the role.\n"
            "- Mention strengths, weaknesses, missed edge cases, strong approaches, and communication quality.\n"
            "- Explicitly mention plagiarism true/false where relevant and that plagiarized answers get score 0.\n"
            "- Mention warning behavior and whether warnings indicate risk.\n"
            "- Give a final recommendation and overall summary.\n"
            "- Return readable markdown with short sections and bullets where helpful."
        )
        msg = client.messages.create(
            model=settings.claude_model,
            max_tokens=3072,
            messages=[{"role": "user", "content": prompt}],
        )
        return (msg.content[0].text if msg.content else "").strip() or "No report generated."
    except Exception as exc:
        logger.exception("Report generation failed: %s", exc)
        return f"Report generation failed: {exc}"


@router.get("/me", response_model=CompanyRead)
def get_my_company(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> CompanyRead:
    company = get_company_for_user(db, current_user)
    return company


@router.post("/candidates", response_model=CandidateRead)
def create_candidate(
    payload: CandidateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> CandidateRead:
    company = get_company_for_user(db, current_user)

    existing_user = db.query(User).filter(User.email == payload.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Candidate email already exists")

    user = User(
        email=payload.email,
        full_name=payload.full_name,
        hashed_password=get_password_hash(payload.password),
        role=role_value(UserRole.CANDIDATE),
    )
    db.add(user)
    db.flush()

    candidate = Candidate(user_id=user.id, company_id=company.id)
    db.add(candidate)
    db.commit()
    db.refresh(candidate)

    return CandidateRead(
        id=candidate.id,
        email=user.email,
        full_name=user.full_name,
        company_id=company.id,
        created_at=candidate.created_at,
    )


@router.get("/candidates", response_model=list[CandidateRead])
def list_candidates(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[CandidateRead]:
    company = get_company_for_user(db, current_user)
    candidates = (
        db.query(Candidate)
        .filter(Candidate.company_id == company.id)
        .order_by(Candidate.created_at.desc())
        .all()
    )
    result: list[CandidateRead] = []
    for c in candidates:
        result.append(
            CandidateRead(
                id=c.id,
                email=c.user.email,
                full_name=c.user.full_name,
                company_id=c.company_id,
                created_at=c.created_at,
            )
        )
    return result


@router.post("/interviews", response_model=InterviewRead)
def create_interview(
    payload: InterviewCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> InterviewRead:
    company = get_company_for_user(db, current_user)
    interview = Interview(
        name=payload.name,
        description=payload.description,
        company_id=company.id,
        follow_order=payload.follow_order,
        shortlist_count=payload.shortlist_count,
        scheduled_start=payload.scheduled_start,
        scheduled_end=payload.scheduled_end,
    )
    db.add(interview)
    db.commit()
    db.refresh(interview)
    return interview


@router.put("/interviews/{interview_id}", response_model=InterviewRead)
def update_interview(
    interview_id: int,
    payload: InterviewUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> InterviewRead:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    for field in ("name", "description", "follow_order", "shortlist_count", "scheduled_start", "scheduled_end"):
        val = getattr(payload, field, None)
        if val is not None:
            setattr(interview, field, val)
    db.commit()
    db.refresh(interview)
    return interview


@router.patch("/interviews/{interview_id}/activate", response_model=InterviewRead)
def activate_interview(
    interview_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> InterviewRead:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    if interview.status not in ("draft",):
        raise HTTPException(status_code=400, detail=f"Cannot activate from status '{interview.status}'")
    rounds = db.query(Round).filter(Round.interview_id == interview.id).all()
    if not rounds:
        raise HTTPException(status_code=400, detail="Add at least one round before activating")
    total_weightage = sum(r.weightage for r in rounds)
    if total_weightage > 0 and abs(total_weightage - 100) > 0.01:
        raise HTTPException(status_code=400, detail=f"Round weightages must sum to 100 (current: {total_weightage})")
    interview.status = "active"
    db.commit()
    db.refresh(interview)
    return interview


@router.patch("/interviews/{interview_id}/end", response_model=InterviewRead)
def end_interview(
    interview_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> InterviewRead:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    if interview.status not in ("active", "in_progress"):
        raise HTTPException(status_code=400, detail=f"Cannot end from status '{interview.status}'")
    interview.status = "completed"
    db.commit()
    db.refresh(interview)
    return interview


@router.patch("/interviews/{interview_id}/terminate", response_model=InterviewRead)
def terminate_interview(
    interview_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> InterviewRead:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    interview.status = "terminated"
    db.commit()
    db.refresh(interview)
    return interview


@router.get("/interviews", response_model=list[InterviewRead])
def list_interviews(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[InterviewRead]:
    company = get_company_for_user(db, current_user)
    interviews = (
        db.query(Interview)
        .filter(Interview.company_id == company.id)
        .order_by(Interview.created_at.desc())
        .all()
    )
    return interviews


@router.post("/interviews/{interview_id}/rounds", response_model=RoundRead)
def add_round_to_interview(
    interview_id: int,
    payload: RoundCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> RoundRead:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")

    round_obj = Round(
        interview_id=interview.id,
        type=payload.type,
        order=payload.order,
        weightage=payload.weightage,
        duration_minutes=payload.duration_minutes,
        config=payload.config.model_dump() if payload.config else None,
    )
    db.add(round_obj)
    db.commit()
    db.refresh(round_obj)
    return round_obj


@router.get("/interviews/{interview_id}/rounds", response_model=list[RoundRead])
def list_rounds_for_interview(
    interview_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[RoundRead]:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")

    rounds = (
        db.query(Round)
        .filter(Round.interview_id == interview.id)
        .order_by(Round.order.asc())
        .all()
    )
    return rounds


@router.get("/interviews/{interview_id}/candidates", response_model=list[EnrolledCandidateRead])
def list_enrolled_candidates(
    interview_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[EnrolledCandidateRead]:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    links = (
        db.query(InterviewCandidate)
        .filter(InterviewCandidate.interview_id == interview.id)
        .all()
    )
    out: list[EnrolledCandidateRead] = []
    for link in links:
        c = db.query(Candidate).filter(Candidate.id == link.candidate_id).first()
        if not c:
            continue
        u = c.user
        ver = db.query(Verification).filter(Verification.candidate_id == c.id).first()
        status = (ver.status if ver else None) or "not_submitted"
        out.append(
            EnrolledCandidateRead(
                id=c.id,
                email=u.email,
                full_name=u.full_name,
                verification_status=status,
            )
        )
    return out


@router.post("/interviews/{interview_id}/candidates")
def enroll_candidates_to_interview(
    interview_id: int,
    candidate_ids: list[int],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")

    candidates = (
        db.query(Candidate)
        .filter(Candidate.company_id == company.id, Candidate.id.in_(candidate_ids))
        .all()
    )
    if not candidates:
        raise HTTPException(status_code=400, detail="No valid candidates to enroll")

    created = 0
    for c in candidates:
        exists = (
            db.query(InterviewCandidate)
            .filter(
                InterviewCandidate.interview_id == interview.id,
                InterviewCandidate.candidate_id == c.id,
            )
            .first()
        )
        if exists:
            continue
        link = InterviewCandidate(interview_id=interview.id, candidate_id=c.id)
        db.add(link)
        created += 1

    db.commit()
    return {"enrolled": created}


@router.delete("/interviews/{interview_id}/candidates/{candidate_id}")
def remove_candidate_from_interview(
    interview_id: int,
    candidate_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    """Remove candidate from interview and delete their data from DB (Candidate + User)."""
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    candidate_obj = (
        db.query(Candidate)
        .filter(Candidate.id == candidate_id, Candidate.company_id == company.id)
        .first()
    )
    if not candidate_obj:
        raise HTTPException(status_code=404, detail="Candidate not found")
    link = (
        db.query(InterviewCandidate)
        .filter(
            InterviewCandidate.interview_id == interview_id,
            InterviewCandidate.candidate_id == candidate_id,
        )
        .first()
    )
    if link:
        db.delete(link)
    db.query(ProctoringEvent).filter(ProctoringEvent.candidate_id == candidate_id).delete()
    db.query(Strike).filter(Strike.candidate_id == candidate_id).delete()
    db.query(Response).filter(Response.candidate_id == candidate_id).delete()
    db.query(Verification).filter(Verification.candidate_id == candidate_id).delete()
    db.query(InterviewCandidate).filter(InterviewCandidate.candidate_id == candidate_id).delete()
    user_id = candidate_obj.user_id
    db.delete(candidate_obj)
    user = db.query(User).filter(User.id == user_id).first()
    if user:
        db.delete(user)
    db.commit()
    return {"ok": True, "message": "Candidate removed and data deleted"}


@router.get("/rounds/{round_id}", response_model=RoundRead)
def get_round(
    round_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> RoundRead:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    return round_obj


@router.patch("/rounds/{round_id}", response_model=RoundRead)
def update_round_config(
    round_id: int,
    payload: RoundConfigUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> RoundRead:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    config = dict(round_obj.config or {})
    if payload.recruiter_requirements is not None:
        config["recruiter_requirements"] = payload.recruiter_requirements
    if payload.resume_summary is not None:
        config["resume_summary"] = payload.resume_summary
    round_obj.config = config
    db.commit()
    db.refresh(round_obj)
    return round_obj


@router.put("/rounds/{round_id}", response_model=RoundRead)
def update_round(
    round_id: int,
    payload: RoundUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> RoundRead:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    for field in ("type", "order", "weightage", "duration_minutes"):
        val = getattr(payload, field, None)
        if val is not None:
            setattr(round_obj, field, val)
    if payload.config is not None:
        round_obj.config = payload.config.model_dump()
    db.commit()
    db.refresh(round_obj)
    return round_obj


@router.delete("/rounds/{round_id}")
def delete_round(
    round_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    q_ids = [q.id for q in db.query(Question).filter(Question.round_id == round_id).all()]
    if q_ids:
        db.query(Response).filter(Response.question_id.in_(q_ids)).delete(synchronize_session=False)
    db.query(Question).filter(Question.round_id == round_id).delete(synchronize_session=False)
    db.query(RoundSession).filter(RoundSession.round_id == round_id).delete(synchronize_session=False)
    db.query(Strike).filter(Strike.round_id == round_id).delete(synchronize_session=False)
    db.query(ProctoringEvent).filter(ProctoringEvent.round_id == round_id).delete(synchronize_session=False)
    db.delete(round_obj)
    db.commit()
    return {"ok": True}


@router.put("/interviews/{interview_id}/rounds/reorder")
def reorder_rounds(
    interview_id: int,
    payload: RoundsReorderRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    rounds = db.query(Round).filter(Round.interview_id == interview.id).all()
    round_map = {r.id: r for r in rounds}
    for idx, rid in enumerate(payload.round_ids):
        if rid in round_map:
            round_map[rid].order = idx + 1
    db.commit()
    return {"ok": True}


@router.post("/rounds/{round_id}/questions", response_model=QuestionRead)
def add_question_to_round(
    round_id: int,
    payload: QuestionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> QuestionRead:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")

    question = Question(
        round_id=round_obj.id,
        content=payload.content,
        type=payload.type,
        difficulty=payload.difficulty,
        domain=payload.domain,
        options=payload.options,
        correct_answer=payload.correct_answer,
        max_score=payload.max_score,
        test_cases=payload.test_cases,
        extra_metadata=payload.extra_metadata,
    )
    db.add(question)
    db.commit()
    db.refresh(question)
    return question


@router.put("/questions/{question_id}", response_model=QuestionRead)
def update_question(
    question_id: int,
    payload: QuestionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> QuestionRead:
    company = get_company_for_user(db, current_user)
    q = (
        db.query(Question)
        .join(Round, Question.round_id == Round.id)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Question.id == question_id, Interview.company_id == company.id)
        .first()
    )
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    for field in ("content", "type", "difficulty", "domain", "options", "correct_answer", "max_score", "test_cases", "extra_metadata"):
        val = getattr(payload, field, None)
        if val is not None:
            setattr(q, field, val)
    db.commit()
    db.refresh(q)
    return q


@router.delete("/questions/{question_id}")
def delete_question(
    question_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    q = (
        db.query(Question)
        .join(Round, Question.round_id == Round.id)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Question.id == question_id, Interview.company_id == company.id)
        .first()
    )
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    db.query(Response).filter(Response.question_id == question_id).delete(synchronize_session=False)
    db.delete(q)
    db.commit()
    return {"ok": True}


@router.patch("/questions/{question_id}/approve")
def approve_question(
    question_id: int,
    approved: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    q = (
        db.query(Question)
        .join(Round, Question.round_id == Round.id)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Question.id == question_id, Interview.company_id == company.id)
        .first()
    )
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    q.approved = approved
    db.commit()
    return {"ok": True, "approved": approved}


@router.get("/rounds/{round_id}/questions", response_model=list[QuestionRead])
def list_round_questions(
    round_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[QuestionRead]:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")

    questions = db.query(Question).filter(Question.round_id == round_obj.id).all()
    return questions


@router.post("/rounds/{round_id}/questions/generate", response_model=list[QuestionRead])
def generate_round_questions(
    round_id: int,
    payload: GenerateQuestionsRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[QuestionRead]:
    try:
        company = get_company_for_user(db, current_user)
        round_obj = (
            db.query(Round)
            .join(Interview, Round.interview_id == Interview.id)
            .filter(Round.id == round_id, Interview.company_id == company.id)
            .first()
        )
        if not round_obj:
            raise HTTPException(status_code=404, detail="Round not found")
        from app.services.question_generator import generate_questions
        count = min(max(1, payload.count), 50)
        generated = generate_questions(
            round_type=round_obj.type,
            count=count,
            difficulty=payload.difficulty,
            domain=payload.domain,
        )
        if not generated:
            raise HTTPException(
                status_code=503,
                detail="No questions were generated. Ensure CLAUDE_API_KEY is set in backend .env and the API is reachable. Check backend logs for details.",
            )
        for g in generated:
            if not g.get("content"):
                continue
            q = Question(
                round_id=round_obj.id,
                content=g["content"],
                type=g.get("type") or "text",
                difficulty=g.get("difficulty"),
                domain=g.get("domain"),
                options=g.get("options"),
                correct_answer=g.get("correct_answer"),
                approved=False,
                test_cases=g.get("test_cases"),
                extra_metadata=g.get("extra_metadata"),
            )
            db.add(q)
        db.commit()
        questions = db.query(Question).filter(Question.round_id == round_obj.id).order_by(Question.id.desc()).limit(count).all()
        return list(reversed(questions))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Generate questions failed: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"Question generation failed: {str(e)}. Ensure CLAUDE_API_KEY is set in backend .env and the database schema is up to date (restart the backend).",
        ) from e


def _parse_question_rows(file: UploadFile) -> list[dict[str, Any]]:
    """Parse CSV/Excel for questions: columns content, type, difficulty, domain (or content, question_type, etc.)."""
    rows = _parse_rows_from_upload(file)
    out: list[dict[str, Any]] = []
    for row in rows:
        content = (row.get("content") or row.get("question") or row.get("question_content") or "").strip()
        if not content:
            continue
        qtype = (row.get("type") or row.get("question_type") or "text").strip() or "text"
        diff = (row.get("difficulty") or "").strip() or None
        dom = (row.get("domain") or "").strip() or None
        out.append({"content": content, "type": qtype, "difficulty": diff, "domain": dom})
    return out


@router.post("/rounds/{round_id}/questions/bulk", response_model=BulkQuestionsResult)
def bulk_add_questions_to_round(
    round_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> BulkQuestionsResult:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    rows = _parse_question_rows(file)
    created = 0
    errors: list[dict] = []
    for i, row in enumerate(rows):
        try:
            q = Question(
                round_id=round_obj.id,
                content=row["content"],
                type=row.get("type") or "text",
                difficulty=row.get("difficulty"),
                domain=row.get("domain"),
            )
            db.add(q)
            db.commit()
            created += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": i + 2, "error": str(e)})
    return BulkQuestionsResult(created=created, failed=len(errors), errors=errors)


def _parse_rows_from_upload(file: UploadFile) -> list[dict[str, Any]]:
    content = file.file.read()
    file.file.seek(0)
    rows: list[dict[str, Any]] = []
    fn = (file.filename or "").lower()
    if fn.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
        for r in reader:
            rows.append(dict(r))
    elif fn.endswith(".xlsx") or fn.endswith(".xls"):
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            return []
        header = [str(c.value or "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2):
            rows.append({header[i]: str(c.value or "").strip() for i, c in enumerate(row) if i < len(header)})
        wb.close()
    return rows


@router.post("/interviews/{interview_id}/candidates/bulk", response_model=BulkCandidateResult)
def bulk_add_candidates_to_interview(
    interview_id: int,
    file: UploadFile = File(...),
    password_column_index: int = 0,
    email_column_index: int = 0,
    name_column_index: int = 1,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> BulkCandidateResult:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")

    rows = _parse_rows_from_upload(file)
    if not rows:
        return BulkCandidateResult(created=0, enrolled=0, failed=0, errors=[{"row": 0, "error": "No rows in file"}])
    keys = list(rows[0].keys()) if rows else []
    def _col(aliases: list[str], default_idx: int) -> str:
        lower_keys = [k.lower().strip() for k in keys]
        for a in aliases:
            for i, lk in enumerate(lower_keys):
                if a == lk or a in lk:
                    return keys[i]
        return keys[default_idx] if default_idx < len(keys) else keys[0]
    email_key = _col(["email", "e-mail", "mail"], email_column_index)
    name_key = _col(["name", "full_name", "full name", "candidate"], name_column_index)
    password_key = _col(["password", "pass", "pwd"], password_column_index)
    created = 0
    enrolled = 0
    errors: list[dict] = []
    for i, row in enumerate(rows):
        row_num = i + 2
        email = (row.get(email_key) or "").strip().lower()
        name = (row.get(name_key) or "").strip()
        password = (row.get(password_key) or "").strip()
        if not email or not password:
            errors.append({"row": row_num, "error": "Missing email or password"})
            continue
        try:
            existing_user = db.query(User).filter(User.email == email).first()
            if existing_user:
                existing_candidate = db.query(Candidate).filter(Candidate.user_id == existing_user.id, Candidate.company_id == company.id).first()
                if existing_candidate:
                    link = (
                        db.query(InterviewCandidate)
                        .filter(InterviewCandidate.interview_id == interview.id, InterviewCandidate.candidate_id == existing_candidate.id)
                        .first()
                    )
                    if not link:
                        db.add(InterviewCandidate(interview_id=interview.id, candidate_id=existing_candidate.id))
                        enrolled += 1
                    db.commit()
                else:
                    errors.append({"row": row_num, "error": f"User exists but not as candidate for this company"})
                continue
            user = User(
                email=email,
                full_name=name or None,
                hashed_password=get_password_hash(password),
                role=role_value(UserRole.CANDIDATE),
            )
            db.add(user)
            db.flush()
            candidate = Candidate(user_id=user.id, company_id=company.id)
            db.add(candidate)
            db.flush()
            db.add(InterviewCandidate(interview_id=interview.id, candidate_id=candidate.id))
            db.commit()
            created += 1
            enrolled += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": row_num, "error": str(e)})
    return BulkCandidateResult(created=created, enrolled=enrolled, failed=len(errors), errors=errors)


# ---- Response views: Interview → Round → Question → Candidate ----
@router.get("/interviews/{interview_id}/rounds/{round_id}/responses")
def list_responses_by_round(
    interview_id: int,
    round_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[ResponseWithCandidate]:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id, Interview.id == interview_id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    responses = (
        db.query(Response)
        .join(Question, Response.question_id == Question.id)
        .filter(Question.round_id == round_id)
        .all()
    )
    result = []
    for r in responses:
        c = db.query(Candidate).filter(Candidate.id == r.candidate_id).first()
        q = db.query(Question).filter(Question.id == r.question_id).first()
        strike = db.query(Strike).filter(
            Strike.interview_id == interview_id,
            Strike.round_id == round_id,
            Strike.candidate_id == r.candidate_id,
        ).first()
        if not c or not q:
            continue
        plagiarism_warning = _has_plagiarism_warning(r.flags)
        result.append(
            ResponseWithCandidate(
                id=r.id,
                candidate_id=c.id,
                candidate_email=c.user.email,
                candidate_name=c.user.full_name,
                question_id=q.id,
                question_content=q.content[:200] + ("..." if len(q.content) > 200 else ""),
                question_type=q.type,
                content=r.content,
                score=r.score,
                effective_score=0 if plagiarism_warning else r.score,
                warning_count=strike.strikes if strike else 0,
                plagiarism_warning=plagiarism_warning,
                plagiarism=(r.flags or {}).get("plagiarism") if r.flags else None,
                cross_plagiarism=(r.flags or {}).get("cross_plagiarism") if r.flags else None,
                created_at=r.created_at,
            )
        )
    return result


@router.get("/interviews/{interview_id}/top-performers", response_model=list[TopPerformerEntry])
def get_top_performers(
    interview_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[TopPerformerEntry]:
    """Rank candidates by weighted score across all rounds."""
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    rankings = _compute_interview_rankings(db, interview)
    return [
        TopPerformerEntry(
            candidate_id=item["candidate_id"],
            candidate_email=item["candidate_email"],
            candidate_name=item["candidate_name"],
            total_weighted_score=item["total_weighted_score"],
            round_scores={f"round_{rid}": details["normalized_score"] for rid, details in item["round_scores"].items()},
            recommendation_status=item.get("recommendation_status"),
        )
        for item in rankings
    ]


@router.post("/interviews/{interview_id}/create-from-shortlisted", response_model=InterviewRead)
def create_from_shortlisted(
    interview_id: int,
    name: str = "Next Round",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> InterviewRead:
    """Create a new interview from the top N shortlisted candidates."""
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")

    top = get_top_performers(interview_id, db, current_user)
    limit = interview.shortlist_count or len(top)
    shortlisted_ids = [t.candidate_id for t in top[:limit]]
    if not shortlisted_ids:
        raise HTTPException(status_code=400, detail="No candidates to shortlist")

    new_interview = Interview(
        name=name,
        company_id=company.id,
        description=f"Created from shortlist of interview #{interview.id}",
    )
    db.add(new_interview)
    db.flush()
    for cid in shortlisted_ids:
        db.add(InterviewCandidate(interview_id=new_interview.id, candidate_id=cid))
    db.commit()
    db.refresh(new_interview)
    return new_interview


@router.post("/interviews/{interview_id}/candidates/{candidate_id}/report")
def generate_candidate_report(
    interview_id: int,
    candidate_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    """Generate an AI report for a candidate's performance across all rounds."""
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    verification = db.query(Verification).filter(Verification.candidate_id == candidate_id).first()
    summary = _build_candidate_performance_summary(db, interview, candidate, verification)
    report_text = _generate_candidate_report_text(interview, summary)
    return {
        "report": report_text,
        "summary": {
            **summary.model_dump(),
            "report": report_text,
        },
    }


@router.get("/rounds/{round_id}/plagiarism-check")
def check_plagiarism_for_round(
    round_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    """Run plagiarism detection across all responses in a round."""
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")

    questions = db.query(Question).filter(Question.round_id == round_id).all()
    from app.services.plagiarism import check_code_plagiarism, check_cross_plagiarism, has_plagiarism_warning
    results = []
    flagged_count = 0
    impacted_candidates: set[int] = set()
    for q in questions:
        responses = db.query(Response).filter(Response.question_id == q.id).all()
        all_codes = [(r.candidate_id, r.content or "") for r in responses if r.content]
        for r in responses:
            if not r.content or not r.content.strip():
                continue
            pl = check_code_plagiarism(r.content, q.id, r.candidate_id)
            others = [(cid, code) for cid, code in all_codes if cid != r.candidate_id]
            cross = check_cross_plagiarism(r.content, q.id, r.candidate_id, others)
            candidate = db.query(Candidate).filter(Candidate.id == r.candidate_id).first()
            if pl.get("warning") or cross.get("warning"):
                flagged_count += 1
            r.flags = {
                **dict(r.flags or {}),
                "plagiarism": pl,
                "cross_plagiarism": cross,
            }
            if has_plagiarism_warning(r.flags):
                r.score = 0
                details = dict(r.grading_details or {})
                details["plagiarism_penalty_applied"] = True
                r.grading_details = details
                r.grading_method = (r.grading_method or "auto_code") + "_plagiarism_zeroed"
                impacted_candidates.add(r.candidate_id)
            results.append({
                "candidate_id": r.candidate_id,
                "candidate_name": candidate.user.full_name if candidate and candidate.user else None,
                "candidate_email": candidate.user.email if candidate and candidate.user else None,
                "question_id": q.id,
                "question_preview": q.content[:120] + ("..." if len(q.content) > 120 else ""),
                "plagiarism": pl,
                "cross_plagiarism": cross,
            })
    for candidate_id in impacted_candidates:
        _recalculate_round_session_score(db, round_id, candidate_id)
    db.commit()
    return {
        "round_id": round_id,
        "total_checks": len(results),
        "flagged_submissions": flagged_count,
        "checks": results,
    }


# ---- Response views: Interview → Candidate → Round → Question ----
@router.get("/interviews/{interview_id}/candidates/{candidate_id}/responses")
def list_responses_by_candidate(
    interview_id: int,
    candidate_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> list[ResponseWithCandidate]:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    link = (
        db.query(InterviewCandidate)
        .filter(
            InterviewCandidate.interview_id == interview_id,
            InterviewCandidate.candidate_id == candidate_id,
        )
        .first()
    )
    if not link:
        raise HTTPException(status_code=404, detail="Candidate not enrolled in this interview")
    round_ids = [ro.id for ro in db.query(Round).filter(Round.interview_id == interview_id).all()]
    question_ids_in_rounds = {q.id for q in db.query(Question).filter(Question.round_id.in_(round_ids)).all()} if round_ids else set()
    responses = db.query(Response).filter(Response.candidate_id == candidate_id, Response.question_id.in_(question_ids_in_rounds)).all()
    result = []
    for r in responses:
        if r.question_id not in question_ids_in_rounds:
            continue
        c = db.query(Candidate).filter(Candidate.id == r.candidate_id).first()
        q = db.query(Question).filter(Question.id == r.question_id).first()
        if not c or not q:
            continue
        result.append(
            ResponseWithCandidate(
                id=r.id,
                candidate_id=c.id,
                candidate_email=c.user.email,
                candidate_name=c.user.full_name,
                question_id=q.id,
                question_content=q.content[:200] + ("..." if len(q.content) > 200 else ""),
                content=r.content,
                score=r.score,
                created_at=r.created_at,
            )
        )
    return result


@router.get("/interviews/{interview_id}/candidates/{candidate_id}/summary", response_model=CandidatePerformanceSummary)
def get_candidate_performance_summary(
    interview_id: int,
    candidate_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> CandidatePerformanceSummary:
    company = get_company_for_user(db, current_user)
    interview = (
        db.query(Interview)
        .filter(Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    link = (
        db.query(InterviewCandidate)
        .filter(
            InterviewCandidate.interview_id == interview_id,
            InterviewCandidate.candidate_id == candidate_id,
        )
        .first()
    )
    if not link:
        raise HTTPException(status_code=404, detail="Candidate not enrolled in this interview")

    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    verification = db.query(Verification).filter(Verification.candidate_id == candidate_id).first()
    summary = _build_candidate_performance_summary(db, interview, candidate, verification)
    summary.report = ""  # Report generated on demand via POST /report to avoid slow AI call blocking view
    return summary


@router.post("/interviews/{interview_id}/rounds/{round_id}/live-start", response_model=LiveInterviewStartResponse)
def start_live_interview(
    interview_id: int,
    round_id: int,
    payload: LiveInterviewStartRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> LiveInterviewStartResponse:
    from app.services.jitsi_jwt_service import generate_jitsi_jwt, get_jitsi_domain, get_jitsi_room_name

    company = get_company_for_user(db, current_user)
    interview = db.query(Interview).filter(Interview.id == interview_id, Interview.company_id == company.id).first()
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    round_obj = db.query(Round).filter(Round.id == round_id, Round.interview_id == interview_id).first()
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    if round_obj.type != "LIVE_INTERVIEW":
        raise HTTPException(status_code=400, detail="This endpoint is for LIVE_INTERVIEW rounds only")
    candidate = db.query(Candidate).filter(Candidate.id == payload.candidate_id).first()
    link = db.query(InterviewCandidate).filter(
        InterviewCandidate.interview_id == interview_id,
        InterviewCandidate.candidate_id == payload.candidate_id,
    ).first()
    if not candidate or not link:
        raise HTTPException(status_code=404, detail="Candidate not found in this interview")

    existing = db.query(RoundSession).filter(
        RoundSession.candidate_id == payload.candidate_id,
        RoundSession.round_id == round_id,
    ).first()
    if existing and existing.status == "submitted":
        raise HTTPException(status_code=400, detail="Live interview already completed for this candidate")

    domain = get_jitsi_domain()
    interviewer_name = company.name or current_user.full_name or current_user.email
    if existing and existing.meeting_room_name:
        full_room = get_jitsi_room_name(existing.meeting_room_name)
        interviewer_jwt = generate_jitsi_jwt(existing.meeting_room_name, interviewer_name, current_user.email, True)
        return LiveInterviewStartResponse(
            session_id=existing.id,
            room_name=full_room,
            meeting_url=existing.meeting_url or f"https://{domain}/{full_room}",
            candidate_name=candidate.user.full_name,
            candidate_email=candidate.user.email,
            jitsi_domain=domain,
            jitsi_jwt=interviewer_jwt,
        )

    base_room = f"Neoverse{round_id}{payload.candidate_id}"
    full_room = get_jitsi_room_name(base_room)
    meeting_url = f"https://{domain}/{full_room}"
    if existing:
        existing.meeting_room_name = base_room
        existing.meeting_url = meeting_url
        if not existing.started_at:
            existing.started_at = datetime.utcnow()
        existing.status = "in_progress"
        session = existing
    else:
        session = RoundSession(
            candidate_id=payload.candidate_id,
            round_id=round_id,
            interview_id=interview_id,
            meeting_room_name=base_room,
            meeting_url=meeting_url,
            started_at=datetime.utcnow(),
            status="in_progress",
        )
        db.add(session)
    db.commit()
    db.refresh(session)

    interviewer_jwt = generate_jitsi_jwt(base_room, interviewer_name, current_user.email, True)
    return LiveInterviewStartResponse(
        session_id=session.id,
        room_name=full_room,
        meeting_url=meeting_url,
        candidate_name=candidate.user.full_name,
        candidate_email=candidate.user.email,
        jitsi_domain=domain,
        jitsi_jwt=interviewer_jwt,
    )


@router.post("/interviews/{interview_id}/rounds/{round_id}/live-score", response_model=LiveInterviewScoreResponse)
def score_live_interview(
    interview_id: int,
    round_id: int,
    payload: LiveInterviewScoreRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> LiveInterviewScoreResponse:
    company = get_company_for_user(db, current_user)
    interview = db.query(Interview).filter(Interview.id == interview_id, Interview.company_id == company.id).first()
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    session = db.query(RoundSession).filter(
        RoundSession.candidate_id == payload.candidate_id,
        RoundSession.round_id == round_id,
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="No live session found for this candidate")
    session.total_score = payload.score
    session.max_possible_score = payload.max_score
    session.notes = payload.notes
    session.status = "submitted"
    session.submitted_at = datetime.utcnow()
    db.commit()
    db.refresh(session)
    return LiveInterviewScoreResponse(
        session_id=session.id,
        candidate_id=payload.candidate_id,
        score=payload.score,
        max_score=payload.max_score,
        notes=payload.notes,
        status=session.status,
    )


@router.post("/interviews/{interview_id}/rounds/{round_id}/live-assist", response_model=LiveAssistResponse)
async def live_interview_ai_assist(
    interview_id: int,
    round_id: int,
    payload: LiveAssistRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> LiveAssistResponse:
    from app.services.interview_ai_service import live_interview_assist

    company = get_company_for_user(db, current_user)
    interview = db.query(Interview).filter(Interview.id == interview_id, Interview.company_id == company.id).first()
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    round_obj = db.query(Round).filter(Round.id == round_id, Round.interview_id == interview_id).first()
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    if round_obj.type != "LIVE_INTERVIEW":
        raise HTTPException(status_code=400, detail="This endpoint is for LIVE_INTERVIEW rounds only")
    result = await live_interview_assist(
        conversation_notes=payload.previous_notes,
        latest_note=payload.note,
        config=round_obj.config or {},
    )
    return LiveAssistResponse(
        evaluation=result.get("evaluation", ""),
        evaluation_rating=result.get("evaluation_rating", "neutral"),
        suggested_questions=result.get("suggested_questions", []),
        tip=result.get("tip", ""),
    )


# ---- GD Session Management ----
@router.post("/interviews/{interview_id}/rounds/{round_id}/gd-session")
def start_gd_session(
    interview_id: int,
    round_id: int,
    topic: str | None = None,
    total_turns: int = 20,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    round_obj = (
        db.query(Round)
        .join(Interview, Round.interview_id == Interview.id)
        .filter(Round.id == round_id, Interview.id == interview_id, Interview.company_id == company.id)
        .first()
    )
    if not round_obj:
        raise HTTPException(status_code=404, detail="Round not found")
    if round_obj.type != "GD":
        raise HTTPException(status_code=400, detail="Not a GD round")

    existing = db.query(GDSession).filter(GDSession.round_id == round_id).first()
    if existing and existing.status not in ("ended",):
        return {"id": existing.id, "topic": existing.topic, "status": existing.status}

    from app.services.gd_moderator import generate_gd_topic
    final_topic = topic or generate_gd_topic((round_obj.config or {}).get("domain"))

    links = db.query(InterviewCandidate).filter(InterviewCandidate.interview_id == interview_id).all()
    participant_ids = [l.candidate_id for l in links]

    session = GDSession(
        round_id=round_id,
        topic=final_topic,
        messages=[],
        participant_ids=participant_ids,
        total_turns=total_turns,
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return {"id": session.id, "topic": session.topic, "status": session.status}


@router.post("/interviews/{interview_id}/rounds/{round_id}/gd-session/end")
def end_gd_session(
    interview_id: int,
    round_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    gd = db.query(GDSession).filter(GDSession.round_id == round_id).first()
    if not gd:
        raise HTTPException(status_code=404, detail="No GD session found")
    gd.status = "ended"

    from app.services.gd_moderator import score_gd_participants
    participant_names = []
    for pid in (gd.participant_ids or []):
        c = db.query(Candidate).filter(Candidate.id == pid).first()
        if c:
            participant_names.append(c.user.full_name or c.user.email)
    scores = score_gd_participants(gd.topic, gd.messages or [], participant_names)

    db.commit()
    return {"status": "ended", "scores": scores}


@router.get("/interviews/{interview_id}/rounds/{round_id}/gd-session/status")
def get_gd_session_status(
    interview_id: int,
    round_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    gd = db.query(GDSession).filter(GDSession.round_id == round_id).first()
    if not gd:
        raise HTTPException(status_code=404, detail="No GD session found")
    return {
        "id": gd.id,
        "topic": gd.topic,
        "status": gd.status,
        "turn_number": gd.turn_number,
        "total_turns": gd.total_turns,
        "message_count": len(gd.messages or []),
    }


# ---- Company Dashboard Stats ----
@router.get("/dashboard")
def company_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_company),
) -> dict:
    company = get_company_for_user(db, current_user)
    total_interviews = db.query(Interview).filter(Interview.company_id == company.id).count()
    active_interviews = db.query(Interview).filter(
        Interview.company_id == company.id,
        Interview.status.in_(["active", "in_progress"]),
    ).count()
    total_candidates = db.query(Candidate).filter(Candidate.company_id == company.id).count()
    return {
        "total_interviews": total_interviews,
        "active_interviews": active_interviews,
        "total_candidates": total_candidates,
        "company_name": company.name,
    }

